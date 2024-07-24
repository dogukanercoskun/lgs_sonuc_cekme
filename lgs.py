from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import time
import pandas as pd
import openpyxl
from tkinter import messagebox
from PyQt5.QtWidgets import QApplication, QFileDialog, QLabel, QLineEdit, QVBoxLayout,QPushButton, QDialog
import sys
from selenium.webdriver.support.ui import Select
import logging
import os

logging.basicConfig(filename='error_log.txt', level=logging.ERROR)
def save_progress(index):
    with open('progress.txt', 'w') as f:
        f.write(str(index))
def load_progress():
    try:
        with open('progress.txt', 'r') as f:
            return int(f.read())
    except FileNotFoundError:
        return 0
def delete_progress():
    try:
        os.remove('progress.txt')
    except FileNotFoundError:
        pass
def dosya_ac():
    global write_file,excel_file,read_sheet_name,write_sheet_name
    ogrenci_dosyası_secme_mesajı()
    app = QApplication([])
    options = QFileDialog.Options()
    excel_file, _ = QFileDialog.getOpenFileName(None, "Dosya Seç", "", "All Files (*)", options=options)
    read_workbook=openpyxl.load_workbook(excel_file)
    read_sheet_name=read_workbook.sheetnames[0]
    sonuc_dosyası_secme_mesajı()
    write_file, _ = QFileDialog.getOpenFileName(None, "Dosya Seç", "", "All Files (*)", options=options)
    write_workbook=openpyxl.load_workbook(write_file)
    write_sheet_name=write_workbook.sheetnames[0]
def tercih_yapmayan_ogrenci_veri():
    birthday=str(bDay)+"."+str(bMonth)+"."+str(bYear)
    data = {
    '0': ["T.C. Kimlik No","Adı Soyadı", "Doğum Tarihi", "Sonuç"],
    '1': [str(id),name, birthday, 'T.C. Kimlik Numaranızı/Doğum Tarihinizi Yanlış Girdiniz veya Tercih Başvurunuz Bulunmamaktadır!']} 
    data_frame = pd.DataFrame(data)  
    tablo_verileri.append(data_frame)
def tablo_verilerini_al():
    tablo_elementleri = driver.find_element(By.XPATH, "//div")
    for tablo_elementi in tablo_elementleri.find_elements(By.TAG_NAME, "table"):
        tablo_icerigi = tablo_elementi.get_attribute('outerHTML')
        data_frame = pd.read_html(tablo_icerigi)[0]
        tablo_verileri.append(data_frame)    
def tablo_verilerini_yaz():
    concatenated_df = pd.concat(tablo_verileri)
    write_dolu_satir_sayisi = read_excel_row_count(write_file, write_sheet_name)
    
    # Eğer write_dolu_satir_sayisi 0 ise, startrow 0 olur, aksi takdirde startrow write_dolu_satir_sayisi olur
    startrow = write_dolu_satir_sayisi if write_dolu_satir_sayisi == 0 else write_dolu_satir_sayisi + 1
    print(startrow)
    
    with pd.ExcelWriter(write_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        concatenated_df.to_excel(writer, sheet_name=write_sheet_name, index=False, startrow=startrow, header=write_dolu_satir_sayisi==0)

    # Clear the list after writing to the file
    tablo_verileri.clear()     
def yeni_sorgu():
    yeni_sorgu_selector=(By.PARTIAL_LINK_TEXT, "Yeni Sorgu")
    if is_element_present(driver,*yeni_sorgu_selector):
         yeni_sorgu=driver.find_element(By.PARTIAL_LINK_TEXT, "Yeni Sorgu")
         yeni_sorgu.click()
    else:
        driver.back()
        time.sleep(1) 
def find_input_by_attribute(driver, attribute, value):
    return driver.find_element(By.CSS_SELECTOR, f'input[{attribute}="{value}"]')       
def giris_elemanları():
    tcInput = find_input_by_attribute(driver, "placeholder", "T.C. Kimlik Numarası")
    okulNoInput = find_input_by_attribute(driver, "placeholder", "Okul Numarası")
    #tcInput=driver.find_element(By.XPATH, '//input[(@id="TC_KIMLIK_NO") or (@id="ADAY_NO") or (@id="TCNO")]')
    #okulNoInput=driver.find_element(By.XPATH, '//input[(@id="GUVENLIKNUMARASI") or (@id="OKULNO")]')
    tcInput.clear()
    tcInput.send_keys(str(id))
    okulNoInput.send_keys(str(okulNo))
    selectDay=Select(driver.find_element(By.NAME, "GUN"))
    selectDay.select_by_value(str(bDay)) 
    selectMonth=Select(driver.find_element(By.NAME, "AY"))
    selectMonth.select_by_value(str(bMonth))
    selectYear=Select(driver.find_element(By.NAME, "YIL"))
    selectYear.select_by_value(str(bYear))
def uyarı_mesaj_guvenlik_metin():
      messagebox.showerror("Hata Mesajı", "Girdiğiniz güvelik kodunda bir hata oluştu! sadece güvelik kodunu elle giriniz ve sonra uyarı mesajındaki tamama basınız")
def bilgi_mesajı():
    messagebox.showinfo("Dikkat", "Güvenlik kodunu giriniz ve sonra uyarı mesajındaki tamama basınız")
def sınava_girmeyen_ogrenci_mesajı():
    messagebox.showinfo("Dikkat", f"{name} adlı öğrenci sınava girmemiştir veya tercih başvusunda bulunmamıştır. Tamam basarak sistemin işlemesine devam ediniz")
def bitis_mesajı():
    messagebox.showinfo("Dikkat", "Öğrencilerin sonuçlarını alma işlemi tamamlanmıştır.")
def ogrenci_dosyası_secme_mesajı():
    messagebox.showinfo("Dikkat", "Lgs sonuçları çekilecek olan öğrencilerin excel dosyasını açılan ekrandan seçiniz")
def sonuc_dosyası_secme_mesajı():
    messagebox.showinfo("Dikkat", "Lgs sonuçlarının kaydedileceği excel dosyasını açılan ekrandan seçiniz")
def is_element_present(driver, by, selector):
   
    try:
        driver.find_element(by, selector)
        return True
    except NoSuchElementException:
        return False
def read_excel_row(file_path, sheet_name, row_index):
    # Excel dosyasını okuyalım
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Belirtilen satırdaki verileri bir değişkene atayalım
    row_data = df.iloc[row_index].values.tolist()

    return row_data
def read_excel_row_count(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    dolu_satir_sayisi = df.shape[0]
    
    return dolu_satir_sayisi
def başarılı_sorgu_sonuc():
    tablo_verilerini_al()
    yeni_sorgu()
def get_input():
    app = QApplication(sys.argv)

    dialog = QDialog()
    dialog.setWindowTitle('LGS SONUÇ URL')

    label = QLabel('Sonuçların alınacağı url giriniz:')
    input_field = QLineEdit()

    def on_ok_clicked():
        global input_value
        input_value = input_field.text()
        dialog.accept()

    ok_button = QPushButton('Tamam')
    ok_button.clicked.connect(on_ok_clicked)

    layout = QVBoxLayout()
    layout.addWidget(label)
    layout.addWidget(input_field)
    layout.addWidget(ok_button)

    dialog.setLayout(layout)
    dialog.setFixedSize(300, 150)

    dialog.exec_()
    return input_value
def sonucları_al():
    global name, bDay, bMonth, bYear, id, okulNo
    start_index = load_progress()
    
    for i in range(start_index, dolu_satir_sayisi):
        try:
            row_index = i  
            row_data = read_excel_row(excel_file, read_sheet_name, row_index)
            name, bDay, bMonth, bYear, id, okulNo = row_data[:6]
            
            if int(bDay) < 10:
                bDay = "0" + str(bDay)
            if int(bMonth) < 10:
                bMonth = "0" + str(bMonth)
            
            giris_elemanları()
            guvenlık_kodu_selector = (By.XPATH, '//input[(@id="GUVENLIKKODU") or (@id="gkodu")]')
            hata_kodu_selector = (By.ID, "hata")
            yeni_sayfada_hata_kodu_selector = (By.XPATH, "//p[@align='center']")
            
            tamam_button=driver.find_element(By.NAME, "Submit")
            
            if is_element_present(driver, *guvenlık_kodu_selector):
                bilgi_mesajı()
                tamam_button.click()
                time.sleep(2)
                
                if is_element_present(driver, *hata_kodu_selector):
                    hatakodutext = driver.find_element(By.XPATH, "//*[@id='hata']")
                    if hatakodutext.text == "Güvenlik Kodunu yanlış girdiniz!":
                        uyarı_mesaj_guvenlik_metin()
                        giris_elemanları()
                        time.sleep(1)
                        tamam_button=driver.find_element(By.NAME, "Submit")
                        tamam_button.click()
                        if is_element_present(driver, *hata_kodu_selector):
                            sınava_girmeyen_ogrenci_mesajı()
                            tercih_yapmayan_ogrenci_veri()
                            continue
                    else:
                        sınava_girmeyen_ogrenci_mesajı()
                        tercih_yapmayan_ogrenci_veri()
                        continue
                
                if is_element_present(driver, *yeni_sayfada_hata_kodu_selector):
                    yeni_hatakodutext = driver.find_element(By.XPATH, '//p[1]')
                    if yeni_hatakodutext.text in ["T.C. Kimlik Numaranızı/Doğum Tarihinizi Yanlış Girdiniz veya Tercih Başvurunuz Bulunmamaktadır!", "T.C. Kimlik Numaranızı veya Doğum Tarihinizi Yanlış Girdiniz!"]:
                        sınava_girmeyen_ogrenci_mesajı()
                        tercih_yapmayan_ogrenci_veri()
                        yeni_sorgu()
                        continue
                    else:
                        yeni_sorgu()
                        giris_elemanları()
                        uyarı_mesaj_guvenlik_metin()
                        tamam_button=driver.find_element(By.NAME, "Submit")
                        tamam_button.click()
                        if is_element_present(driver, *yeni_sayfada_hata_kodu_selector):
                            sınava_girmeyen_ogrenci_mesajı()
                            tercih_yapmayan_ogrenci_veri()
                            yeni_sorgu()
                            continue
                
                başarılı_sorgu_sonuc()
            
            else:
                tamam_button.click()
                time.sleep(2)
                if is_element_present(driver, *hata_kodu_selector):
                    sınava_girmeyen_ogrenci_mesajı()
                    tercih_yapmayan_ogrenci_veri()
                    continue
                if is_element_present(driver, *yeni_sayfada_hata_kodu_selector):
                    yeni_sorgu()
                    sınava_girmeyen_ogrenci_mesajı()
                    tercih_yapmayan_ogrenci_veri()
                    continue
                başarılı_sorgu_sonuc()
            
            save_progress(i + 1)  # İlerlemeyi her adımda kaydet
            
            if (i + 1) % 5 == 0:
                tablo_verilerini_yaz()
                
        except Exception as e:
            logging.error(f"Hata oluştu (öğrenci indeksi {i}): {str(e)}")
            save_progress(i)  # Hata durumunda son başarılı konumu kaydet
            tablo_verilerini_yaz()  # Hata durumunda mevcut verileri kaydet
            raise  # Hatayı yeniden fırlat
    
    tablo_verilerini_yaz()  # Son kez kaydet

url=get_input()
read_sheet_name: any
write_sheet_name:any  
write_file:any
excel_file:any
dosya_ac()
dolu_satir_sayisi=read_excel_row_count(excel_file, read_sheet_name)
tablo_verileri=[]
name:any
bDay:any
bMonth:any
bYear:any
id:any
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# driver.maximize_window()
driver.get(url)
time.sleep(2)

try:
    sonucları_al()
    delete_progress()
except Exception as e:
    logging.error(f"Hata oluştu: {str(e)}")

time.sleep(1)
bitis_mesajı()
driver.quit()
